import os
import openpyxl
from utils.point_double_variant import APoint, ADouble, variants

obwod_przekroju = []

def open_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(parent_dir, "data.xlsm")
    sheet_name = "Obliczenia i dane"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"{e}")

    

def draw_lines(model_space, sheet):
    
    for row in range(271, 276):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        obwod_przekroju.append(line)   
        
        print("Linia katownika PD narysowana")   
    
    

def draw_arcs(model_space, sheet):
    
    for row in range(14, 17):
        row_data = []
        for col in range(11, 18):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)         
    
        s = APoint(row_data[0], row_data[1])
        arc = model_space.AddArc(s, row_data[2], row_data[3], row_data[4])
        arc.Layer = row_data[5]
        arc.LinetypeScale = row_data[6]
        
        obwod_przekroju.append(arc) 
        
        print("Łuk katownika PD narysowany")
    
    

def draw_katownik_pasa_dolnego(doc):
    
    model_space = doc.ModelSpace
    
    sheet = open_excel()
    
    draw_lines(model_space, sheet)
    draw_arcs(model_space, sheet)
    