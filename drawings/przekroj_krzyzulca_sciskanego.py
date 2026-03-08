import os
import openpyxl
from utils.point_double_variant import APoint, ADouble, variants

obwod_przekroju1 = []
obwod_przekroju2 = []

def open_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(parent_dir, "data.xlsm")
    sheet_name = "Obliczenia i dane"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"{e}")
        
        
def draw_axes(model_space, sheet):
    
    for row in range(131, 134):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia narysowana") 

    

def draw_lines1(model_space, sheet):
    
    for row in range(119, 125):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        obwod_przekroju1.append(line)   
        
        print("Linia przekroju KŚ narysowana")   
        
        
def draw_lines2(model_space, sheet):
    
    for row in range(125, 131):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        obwod_przekroju2.append(line)   
        
        print("Linia przekroju KŚ narysowana")   
    
    

def draw_arcs1(model_space, sheet):
    
    for row in range(96, 99):
        row_data = []
        for col in range(32, 39):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)         
    
        s = APoint(row_data[0], row_data[1])
        arc = model_space.AddArc(s, row_data[2], row_data[3], row_data[4])
        arc.Layer = row_data[5]
        arc.LinetypeScale = row_data[6]
        
        obwod_przekroju1.append(arc) 
        
        print("Łuk przekroju KŚ narysowany")
        
        
def draw_arcs2(model_space, sheet):
    
    for row in range(99, 102):
        row_data = []
        for col in range(32, 39):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)         
    
        s = APoint(row_data[0], row_data[1])
        arc = model_space.AddArc(s, row_data[2], row_data[3], row_data[4])
        arc.Layer = row_data[5]
        arc.LinetypeScale = row_data[6]
        
        obwod_przekroju2.append(arc) 
        
        print("Łuk przekroju KŚ narysowany")
        

def draw_hatch1(model_space, sheet):
    
    pattern = str(sheet.cell(104, 32).value) 
    scale = float(sheet.cell(104, 33).value) 
    angle = float(sheet.cell(104, 34).value) 
    layer = str(sheet.cell(104, 35).value)   
    
    outer = variants(obwod_przekroju1)
    # Tworzenie zakreskowania (Hatch)
    hatch = model_space.AddHatch(0, pattern, True)  # ANSI31 – standardowy wzór kreskowania

    hatch.PatternScale = scale
    hatch.PatternAngle = angle
    hatch.Layer = layer
    # Dodanie obiektów do hatcha jako pętla graniczna
    hatch.AppendOuterLoop(outer)
    
    
def draw_hatch2(model_space, sheet):
    
    pattern = str(sheet.cell(104, 32).value) 
    scale = float(sheet.cell(104, 33).value) 
    angle = float(sheet.cell(104, 34).value) 
    layer = str(sheet.cell(104, 35).value)   
    
    outer = variants(obwod_przekroju2)
    # Tworzenie zakreskowania (Hatch)
    hatch = model_space.AddHatch(0, pattern, True)  # ANSI31 – standardowy wzór kreskowania

    hatch.PatternScale = scale
    hatch.PatternAngle = angle
    hatch.Layer = layer
    # Dodanie obiektów do hatcha jako pętla graniczna
    hatch.AppendOuterLoop(outer)
    
    

def draw_przekroj_krzyzulca_sciskanego(doc):
    
    model_space = doc.ModelSpace
    
    sheet = open_excel()
    
    draw_lines1(model_space, sheet)
    draw_arcs1(model_space, sheet)
    draw_hatch1(model_space, sheet)
    
    draw_lines2(model_space, sheet)
    draw_arcs2(model_space, sheet)
    draw_hatch2(model_space, sheet)
    
    draw_axes(model_space, sheet)
    