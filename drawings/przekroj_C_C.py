import os
import openpyxl
from utils.point_double_variant import APoint, ADouble, variants

hatch_areas = {}


def open_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(parent_dir, "data.xlsm")
    sheet_name = "Obliczenia i dane"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"{e}")

    

def draw_lines(model_space, sheet):
    
    for row in range(522, 556):
        row_data = []
        for col in range(14, 21):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        if row_data[6] != None:
            main_key = row_data[4]
            if main_key not in hatch_areas:
                hatch_areas[main_key] = {}
                
            key = int(row_data[6])
            if key not in hatch_areas[main_key]:
                hatch_areas[main_key][key] = []
            hatch_areas[main_key][key].append(line)

        # obwod_przekroju1.append(line)   
        
        print("Linia przekroju KŚ narysowana")   
        

def draw_circle(model_space, sheet):
    
    for row in range(492, 496):
        row_data = []
        for col in range(22, 27):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        radius = row_data[2]
        circle = model_space.AddCircle(p1, radius)
        circle.Layer = row_data[3]
        circle.LinetypeScale = row_data[4]
        
        print("Okrąg narysowany")
        
        

def draw_hatch(model_space, sheet):
    
    
    for main_key in hatch_areas:
        
        row = 539
        
        if main_key == str(sheet.cell(540, 25).value):
            row += 1
            
        pattern = str(sheet.cell(row, 22).value) 
        scale = float(sheet.cell(row, 23).value) 
        angle = float(sheet.cell(row, 24).value) 
        layer = str(sheet.cell(row, 25).value)   
        
        for key in hatch_areas[main_key]:            
            outer = variants(hatch_areas[main_key][key])
        
            # Tworzenie zakreskowania (Hatch)
            hatch = model_space.AddHatch(0, pattern, True)  # ANSI31 – standardowy wzór kreskowania

            hatch.PatternScale = scale
            hatch.PatternAngle = angle
            hatch.Layer = layer
            # Dodanie obiektów do hatcha jako pętla graniczna
            hatch.AppendOuterLoop(outer)


def draw_arcs(model_space, sheet):
    
    for row in range(523, 531):
        row_data = []
        for col in range(22, 30):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)         
    
        s = APoint(row_data[0], row_data[1])
        arc = model_space.AddArc(s, row_data[2], row_data[3], row_data[4])
        arc.Layer = row_data[5]
        arc.LinetypeScale = row_data[6]
        
        if row_data[7] != None:
            main_key = row_data[5]
            if main_key not in hatch_areas:
                hatch_areas[main_key] = {}
                
            key = int(row_data[7])
            if key not in hatch_areas[main_key]:
                hatch_areas[main_key][key] = []
            hatch_areas[main_key][key].append(arc)
        
        print("Łuk przekroju C-C narysowany")
    
    

def draw_przekrojCC(doc):
    
    model_space = doc.ModelSpace
    
    sheet = open_excel()
    
    draw_lines(model_space, sheet)
    draw_arcs(model_space, sheet)
    # draw_circle(model_space, sheet)
    draw_hatch(model_space, sheet)

    