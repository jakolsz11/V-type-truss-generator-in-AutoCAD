import os
import openpyxl
from utils.point_double_variant import APoint, ADouble, variants

obwod_przekroju = []

def open_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(parent_dir, "data.xlsm")
    sheet_name = "Obliczenia i dane"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"{e}")
        
        
def draw_axes(model_space, sheet):
    
    for row in range(66, 69):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia narysowana") 

    

def draw_lines(model_space, sheet):
    
    for row in range(54, 66):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        obwod_przekroju.append(line)   
        
        print("Linia narysowana")   
    
    

def draw_arcs(model_space, sheet):
    
    for row in range(10, 14):
        row_data = []
        for col in range(11, 18):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)         
    
        s = APoint(row_data[0], row_data[1])
        arc = model_space.AddArc(s, row_data[2], row_data[3], row_data[4])
        arc.Layer = row_data[5]
        arc.LinetypeScale = row_data[6]
        
        obwod_przekroju.append(arc) 
        
        print("Łuk narysowany")
        

def draw_hatch(model_space):
    
    outer = variants(obwod_przekroju)
    # Tworzenie zakreskowania (Hatch)
    hatch = model_space.AddHatch(0, "ANSI31", True)  # ANSI31 – standardowy wzór kreskowania

    hatch.PatternScale = 2
    hatch.Layer = "Gradzinowanie"
    # Dodanie obiektów do hatcha jako pętla graniczna
    hatch.AppendOuterLoop(outer) 
    
    

def draw_przekroj_pasa_dolnego(doc):
    
    model_space = doc.ModelSpace
    
    sheet = open_excel()
    
    draw_lines(model_space, sheet)
    draw_arcs(model_space, sheet)
    draw_hatch(model_space)
    draw_axes(model_space, sheet)
    