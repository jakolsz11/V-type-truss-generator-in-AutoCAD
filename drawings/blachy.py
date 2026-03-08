import os
import openpyxl
from utils.point_double_variant import APoint, ADouble, variants

def open_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(parent_dir, "data.xlsm")
    sheet_name = "Obliczenia i dane"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"{e}")

    

def draw_blacha_wezlowa(model_space, sheet):
    
    for row in range(144, 149):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia blachy wezlowej narysowana")   
        
        
def draw_zeberko_podporowe(model_space, sheet):
    
    for row in range(149, 153):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia zeberka podporowego narysowana")   
        
        
    
def draw_blacha_wezla_posredniego_nr_1(model_space, sheet):
    
    for row in range(153, 158):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia blachy wezla posredniego nr 1 narysowana")   
        
        
def draw_blacha_wezla_posredniego_nr_2(model_space, sheet):
    
    for row in range(158, 163):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia blachy wezla posredniego nr 2 narysowana")   
    
    
def draw_blachy(doc):
    
    model_space = doc.ModelSpace
    
    sheet = open_excel()
    
    draw_blacha_wezlowa(model_space, sheet)
    draw_zeberko_podporowe(model_space, sheet)
    draw_blacha_wezla_posredniego_nr_1(model_space, sheet)
    draw_blacha_wezla_posredniego_nr_2(model_space, sheet)