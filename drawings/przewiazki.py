import os
import openpyxl
from utils.point_double_variant import APoint, ADouble, variants

def open_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(parent_dir, "data.xlsm")
    sheet_name = "Obliczenia i dane"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        return sheet
    except Exception as e:
        print(f"{e}")

    

def draw_przewiazki_krzyzulca_rozciaganego(model_space, sheet):
    
    number = int(sheet.cell(163, 10).value)
    
    for row in range(163, 163+number*9):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia przewiazki KR narysowana")   
        
        
def draw_przewiazki_krzyzulca_sciskanego(model_space, sheet):
    
    number = int(sheet.cell(217, 10).value)
    
    for row in range(217, 217+number*9):
        row_data = []
        for col in range(4, 10):
            one_data = sheet.cell(row, col).value      
            
            row_data.append(one_data)      
    
        p1 = APoint(row_data[0], row_data[1])
        p2 = APoint(row_data[2], row_data[3])
        line = model_space.AddLine(p1, p2)
        line.Layer = row_data[4]
        line.LinetypeScale = row_data[5]
        
        print("Linia przewiazki KS narysowana")   
        
    
def draw_przewiazki(doc):
    
    model_space = doc.ModelSpace
    
    sheet = open_excel()
    
    draw_przewiazki_krzyzulca_rozciaganego(model_space, sheet)
    draw_przewiazki_krzyzulca_sciskanego(model_space, sheet)
